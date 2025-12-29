import argparse
import os

from openpyxl import load_workbook
from dotenv import load_dotenv
from gspread import Client, service_account, Spreadsheet

load_dotenv()


def client_init_json() -> Client:
    """Создание клиента для работы с Google Sheets."""
    return service_account(filename="google-api-key.json")


def get_table_by_id(client: Client, table_url) -> Spreadsheet:
    """Получение таблицы из Google Sheets по ID таблицы."""
    return client.open_by_key(table_url)


def main(worksheet_name: str, goal_actions_columns: list[str] | None = None) -> None:
    table_id = os.getenv("GOOGLE_TABLE_ID", "")

    columns_mapping = {
        "Дата визита": "Дата",
        "UTM Source": "utm_source",
        "UTM Medium": "utm_medium",
        "UTM Campaign": "utm_campaign",
        "UTM Term": "UTM-Term",
        "Визиты": "visits",
        "Отказы": "bounceRate",
        "Глубина просмотра": "pageDepth",
        "Время на сайте": "avgVisitDurationSeconds",
        "Роботность": "robotPercentage",
    }

    headers_dict = {
        "Дата": "",
        "utm_source": "",
        "utm_medium": "",
        "utm_campaign": "",
        "UTM-Term": "",
        "visits": "",
        "bounceRate": "",
        "pageDepth": "",
        "avgVisitDurationSeconds": "",
        "robotPercentage": "",
        "GoalActions": "",
    }

    google_client = client_init_json()
    google_table = get_table_by_id(google_client, table_id)
    worksheet = google_table.worksheet(worksheet_name)
    
    # Получаем заголовки из первой строки таблицы для определения колонок
    headers = worksheet.row_values(1)
    # Создаем маппинг: ключ -> номер колонки (буква колонки)
    key_to_column = {}
    for idx, header in enumerate(headers):
        if header in headers_dict:
            # Преобразуем индекс (0-based) в букву колонки (A=1, B=2, ...)
            column_letter = chr(ord('A') + idx)
            key_to_column[header] = column_letter
    
    # Определяем, с какой строки начинать запись (последняя заполненная строка + 1)
    # Игнорируем столбцы A (индекс 0) и L (индекс 11), так как там формулы
    start_row = 2
    all_values = worksheet.get_all_values()

    if all_values:
        # Ищем последнюю строку с данными, исключая столбцы A и L
        last_row_with_data = 1  # Начинаем с заголовка
        for row_idx, row in enumerate(all_values[1:], start=2):  # Пропускаем заголовок (строка 1)
            # Проверяем, есть ли данные в столбцах кроме A (индекс 0) и L (индекс 11)
            has_data = False
            for col_idx, cell_value in enumerate(row):
                if col_idx not in [0, 11]:  # Игнорируем столбцы A и L
                    if cell_value and str(cell_value).strip():
                        has_data = True
                        break
            if has_data:
                last_row_with_data = row_idx
        start_row = last_row_with_data + 1
    
    # Собираем только новые данные из файлов
    new_data = []

    for file in os.listdir("tables"):
        if file.endswith(".success") or os.path.exists(os.path.join("tables", f"{file}.success")):
            continue

        xlsx_file = load_workbook(os.path.join("tables", file))
        date = xlsx_file["Отчет"]["A1"].value.strip().split()[-1]
        letter_to_column = {
            cell.column_letter: cell.value
            for cell in xlsx_file["Отчет"][5]
            if cell.value in columns_mapping
        }
        
        # Создаем маппинг для колонок GoalActions, если они указаны
        goal_actions_letter_to_column = {}
        if goal_actions_columns:
            goal_actions_letter_to_column = {
                cell.column_letter: cell.value
                for cell in xlsx_file["Отчет"][5]
                if cell.value in goal_actions_columns
            }

        for row in xlsx_file["Отчет"].iter_rows(min_row=7):
            new_row = {**headers_dict.copy(), "Дата": date}
            goal_actions_sum: float = 0.0

            for cell in row:
                if cell.column_letter in goal_actions_letter_to_column:
                    value = cell.value
                    if value is not None:
                        if isinstance(value, (int, float)):
                            goal_actions_sum += float(value)
                        if isinstance(value, str):
                            goal_actions_sum += float(value.replace(',', '.'))

                if cell.column_letter in letter_to_column:
                    new_row[columns_mapping[letter_to_column[cell.column_letter]]] = cell.value

            new_row["GoalActions"] = goal_actions_sum if goal_actions_sum > 0 else ""
            new_data.append(new_row)

        with open(f"tables/{file}.success", "x") as _:
            continue

    # Записываем данные в соответствующие колонки
    if new_data:
        if not key_to_column:
            print("Предупреждение: Заголовки не найдены в таблице. Убедитесь, что первая строка содержит заголовки.")
            return
        
        # Для каждого ключа собираем все значения для записи в колонку
        updates = []
        for key, column_letter in key_to_column.items():
            values = [[row.get(key, "")] for row in new_data]
            # Записываем колонку, даже если есть хотя бы одно непустое значение
            if any(str(v[0]).strip() if v[0] is not None else "" for v in values):
                range_name = f"{column_letter}{start_row}:{column_letter}{start_row + len(new_data) - 1}"
                updates.append({
                    'range': range_name,
                    'values': values
                })
        
        # Выполняем все обновления батчем
        if updates:
            worksheet.batch_update(updates)
            
            # Форматируем колонки G и J (bounceRate и robotPercentage) как проценты
            percentage_columns = ["bounceRate", "robotPercentage"]
            for key in percentage_columns:
                if key in key_to_column:
                    column_letter = key_to_column[key]
                    # Форматируем диапазон от start_row до последней строки с данными
                    format_range = f"{column_letter}{start_row}:{column_letter}{start_row + len(new_data) - 1}"
                    worksheet.format(format_range, {
                        "numberFormat": {
                            "type": "PERCENT",
                            "pattern": "0.00%"
                        }
                    })



if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Загрузка xlsx файлов в Google таблицы")
    parser.add_argument(
        "--worksheet",
        "-w",
        type=str,
        required=True,
        help="Имя вкладки в Google таблице для загрузки данных",
    )
    parser.add_argument(
        "--goal-actions-columns",
        "-g",
        type=str,
        nargs="+",
        default=None,
        help="Названия колонок в xlsx файле, значения которых будут суммироваться в колонку GoalActions (можно указать несколько через пробел)",
    )
    args = parser.parse_args()
    main(args.worksheet, args.goal_actions_columns)
